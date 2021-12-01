import json
import os

from openpyxl.styles import PatternFill
from seleniumwire import webdriver
import time
from PIL import Image
import openpyxl
import requests
import random
from bs4 import BeautifulSoup

headers = {
    "user-agent": "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.164 Safari/537.36 OPR/77.0.4054.277"
}

count = 2
BRAND_KEY = 0
http_proxy = {'https': 'http://Selyavmore:O7b0ZcF@185.33.85.152:45785',
'http': 'http://Selyavmore:O7b0ZcF@185.33.85.152:45785'
              }
JSON_URL = ""
number = ""
price = ""
href = ""
desc = ""
img = ""
brand = ""
URL_LOWES = ""

def brands(brand):
    try:
        global BRAND_KEY
        wwb = openpyxl.reader.excel.load_workbook(filename="data/brands.xlsx")
        wwb.active = 0
        sheet = wwb.active
        maxrows = sheet.max_row
        for i in range(1, maxrows + 1):
            if sheet["A" + str(i)].value in brand:
                BRAND_KEY += 1
                print(BRAND_KEY)
        wwb.close()
    except Exception:
        BRAND_KEY = 0


def NewUrl():
    global JSON_URL
    producturl = f"https://www.lowes.com/pd/{number}/productdetail/{str(random.randint(1, 1))}/Guest"
    print(producturl)
    JSON_URL = producturl


def openbook(site_price, site_href, site_title, site_src, site_brand, sales_rank_file, url_file, title_file,
             price_file):
    global count, BRAND_KEY
    if count == 2:
        workbook1 = openpyxl.Workbook()
        print(site_price)
        ws = workbook1.active
        ws._images = []
        for rows in ws:
            for cell in rows:
                cell.fill = PatternFill(fill_type=None)
                cell.value = None
        ws['A1'] = 'Brand'
        ws['B1'] = 'Sales Rank: Current'
        ws['C1'] = 'URL: Amazon'
        ws['D1'] = 'Title'
        ws['H1'] = 'Разница'
        ws['G1'] = 'Buy Box: Current'
        ws['I1'] = 'цена на сайте'
        ws['J1'] = 'картинка'
        ws['K1'] = 'ссылка на сайте'
    else:
        workbook1 = openpyxl.reader.excel.load_workbook(filename="data/result.xlsx")
        print(site_price)
        ws = workbook1.active
    ws['A' + str(count)] = site_brand
    ws['B' + str(count)] = sales_rank_file
    ws['C' + str(count)].hyperlink = url_file
    ws['I' + str(count)] = site_price
    ws['K' + str(count)].hyperlink = URL_LOWES
    ws['L' + str(count)] = site_title
    ws['D' + str(count)] = title_file
    ws['H' + str(count)] = price_file / site_price
    ws['G' + str(count)] = price_file
    print("Данные успешно")
    brands(site_brand)
    if BRAND_KEY > 0:
        ws['A' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['B' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['C' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['D' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['E' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['F' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['G' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['H' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['I' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['J' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['K' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
        ws['L' + str(count)].fill = PatternFill(start_color="FEDBCA", end_color="FEDBCA", fill_type="solid")
    else:
        print("Не покрасил")
    response = requests.get(site_src, headers=headers, proxies=http_proxy)
    file = open("image.jpg", "wb")
    file.write(response.content)
    file.close()
    width = 80
    height = 80
    print("Удачно")

    img = Image.open('image.jpg')
    img = img.resize((width, height))
    img.save('image.jpg')
    img = openpyxl.drawing.image.Image('image.jpg')
    ws.add_image(img, 'J' + str(count))
    rd = ws.row_dimensions[count]
    rd.height = 80
    workbook1.save("data/result.xlsx")
    print("Запись в таблицу!")
    count += 1


def LoadInfo():
    try:
        global udachno, JSON_URL, number, price, href, desc, img, brand
        req = requests.get(JSON_URL, headers=headers, proxies=http_proxy)
        data = req.json()
        json_data = json.loads(req.text)
        price = json_data['productDetails'][number]["price"]['analyticsData']['sellingPrice']
        href = url
        desc = data.get("productDetails").get(number).get("product").get("description")
        img = data.get("productDetails").get(number).get("product").get("epc").get("additionalImages")[0].get("baseUrl")
        brand = data.get("productDetails").get(number).get("product").get("brand")
        print(price, href, desc, img, brand)
    except Exception:
        global neudacha
        print("Пустой json")
        neudacha += 1
        print(f"Неудачно:{neudacha}")
        NewUrl()
        neudacha = 0
        LoadInfo()


def Search(searchstring):
    global JSON_URL, number, URL_LOWES
    options = webdriver.ChromeOptions()
    url = "https://www.lowes.com/"
    options.headless = True
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36 OPR/78.0.4093.184")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(
        "accept=text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9")
    proxy_options = {
        "proxy": {
            "https": "https://Selyavmore:O7b0ZcF@185.33.85.152:45785",
            "http": "http://Selyavmore:O7b0ZcF@185.33.85.152:45785",
            "no_proxy": "localhost,127.0.0.1"
        }
    }
    cwd = os.getcwd()
    driver = webdriver.Chrome(
        executable_path=cwd + "\\chromedriver\\chromedriver.exe",
        seleniumwire_options=proxy_options,
        options=options
    )
    driver.implicitly_wait(10)

    try:
        driver.get(url=url)
        time.sleep(5)
        print("Site has been loaded!")
        search = driver.find_element_by_id("search-query")
        search.clear()
        print(f"UPC:{searchstring}")
        search.send_keys(searchstring)
        time.sleep(5)
        search_button = driver.find_element_by_class_name("sb-search-icon").click()
        time.sleep(5)
        print(driver.current_url)
        URL_LOWES = driver.current_url
        if "search" in driver.current_url:
            JSON_URL = ""
            print("no product")
        else:
            number = driver.current_url.rsplit('/', 1)[-1]
            producturl = f"https://www.lowes.com/pd/{number}/productdetail/{str(random.randint(1, 1))}/Guest"
            print(producturl)
            JSON_URL = producturl
    except Exception as ex:
        print(ex)
    finally:
        print("Closing...")
        driver.close()
        driver.quit()


def func(searchstring, sales_rank, url_file, title, price_file):
    try:
        Search(searchstring)
        if JSON_URL == "":
            print("Товара нет!")
        else:
            LoadInfo()
            global price
            if price < price_file / 2:
                print("Подходит цена!")
                try:
                    openbook(price, href, desc, img, brand,
                             sales_rank, url_file, title, price_file)
                except Exception:
                    print("Ошибка")
            else:
                print("Не подходит")

    except Exception:
        print("Неудачно")


wb = openpyxl.reader.excel.load_workbook(filename="data/rei.com.xlsx")
print(wb.sheetnames)
wb.active = 0
sheet = wb.active
maxrows = sheet.max_row
for i in range(2, maxrows):
    print("Строка : " + str(i))
    numbers = sheet['G' + str(i)].value
    price = sheet['D' + str(i)].value
    sales = sheet['C' + str(i)].value
    url = sheet['E' + str(i)].value
    title = sheet['F' + str(i)].value
    if "," in str(numbers):
        list_of_numbers = [int(s) for s in numbers.split(', ')]
        length = len(list_of_numbers)
        print("Количество кодов: " + str(length))
        for i in range(1, length + 1):
            func(str(list_of_numbers[i - 1]), sales, url, title, price)
    else:
        if numbers == "":
            print("Ничего")
        else:
            func(str(numbers), sales, url, title, price)
print("Работа завершена!")
